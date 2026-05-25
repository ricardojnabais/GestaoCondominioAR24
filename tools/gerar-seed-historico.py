"""
Gera seed-historico.json a partir do backup Excel.
Estrutura: tenants, rubricas, receipts, pagamentosDespesa, planos, prestacoes, orcamentos, meta.
"""
import openpyxl
import json
from datetime import datetime
import os, sys

EXCEL = "/mnt/user-data/uploads/Contas_Condominio_2023_2026_-_actualizado_a_25-05-2026.xlsx"
OUT = "/home/claude/GestaoCondominioAR24/data/seed-historico.json"
os.makedirs(os.path.dirname(OUT), exist_ok=True)

wb = openpyxl.load_workbook(EXCEL, data_only=True)

# ──────────────────────────────────────────────────────────
# MAPEAMENTOS
# ──────────────────────────────────────────────────────────

# Coluna no Excel (0-based) → tenantId
# Excel: posição 0 = label "Fracções"
#        posição 1 = R/C.Dto, 2 = R/C.Esq, 3 = 1º Dto, ...
COL_TO_TENANT = {
    1: 'cond_02',   # R/C Dto - Filipe Solha (91‰) - ADMIN
    2: 'cond_01',   # R/C Esq - João Vaz (79‰)
    3: 'cond_04',   # 1º Dto - Sílvia Gonçalves (87‰) ⚠ em atraso
    4: 'cond_03',   # 1º Esq - Leonel Venâncio (119‰)
    5: 'cond_06',   # 2º Dto - António Figueiredo (88‰)
    6: 'cond_05',   # 2º Esq - Ricardo Cordeiro (121‰)
    7: 'cond_08',   # 3º Dto - Lurdes Serafim (88‰)
    8: 'cond_07',   # 3º Esq - Nuno Silva (115‰)
    9: 'cond_10',   # 4º Dto - Vitor Barata (87‰)
    10: 'cond_09',  # 4º Esq - José Carlos Monteiro (125‰)
}

# Tenants completos (com rentByYear de cada ano)
TENANTS = [
    {'id': 'cond_01', 'name': 'João Vaz',           'fraction': 'R/C Esquerdo', 'permilage': 79,
     'nif': '129465380', 'email': 'joaovaz@example.com',          'isAdmin': False,
     'rentByYear': {'2021': 2800, '2022': 2800, '2023': 3200, '2024': 3200, '2025': 3200, '2026': 3200}},
    {'id': 'cond_02', 'name': 'Filipe Solha',       'fraction': 'R/C Direito',  'permilage': 91,
     'nif': '219481342', 'email': 'filipesolha@gmail.com',         'isAdmin': True,
     'rentByYear': {'2021': 3200, '2022': 3200, '2023': 3700, '2024': 3700, '2025': 3700, '2026': 3700}},
    {'id': 'cond_03', 'name': 'Leonel Venâncio',    'fraction': '1.º Esquerdo', 'permilage': 119,
     'nif': '209959746', 'email': 'leonelvenancio@example.com',    'isAdmin': False,
     'rentByYear': {'2021': 4200, '2022': 4200, '2023': 4800, '2024': 4800, '2025': 4800, '2026': 4800}},
    {'id': 'cond_04', 'name': 'Sílvia Gonçalves',   'fraction': '1.º Direito',  'permilage': 87,
     'nif': '195084381', 'email': 'silviagoncalves@example.com',   'isAdmin': False,
     'rentByYear': {'2021': 3100, '2022': 3100, '2023': 3600, '2024': 3600, '2025': 3600, '2026': 3600}},
    {'id': 'cond_05', 'name': 'Ricardo Cordeiro',   'fraction': '2.º Esquerdo', 'permilage': 121,
     'nif': '214490041', 'email': 'ricardojnabais@gmail.com',      'isAdmin': True,
     'rentByYear': {'2021': 4300, '2022': 4300, '2023': 4900, '2024': 4900, '2025': 4900, '2026': 4900}},
    {'id': 'cond_06', 'name': 'António Figueiredo', 'fraction': '2.º Direito',  'permilage': 88,
     'nif': '101744137', 'email': 'antoniofigueiredo@example.com', 'isAdmin': False,
     'rentByYear': {'2021': 3100, '2022': 3100, '2023': 3600, '2024': 3600, '2025': 3600, '2026': 3600}},
    {'id': 'cond_07', 'name': 'Nuno Silva',         'fraction': '3.º Esquerdo', 'permilage': 115,
     'nif': '195611004', 'email': 'nunosilva@example.com',         'isAdmin': False,
     'rentByYear': {'2021': 4100, '2022': 4100, '2023': 4700, '2024': 4700, '2025': 4700, '2026': 4700}},
    {'id': 'cond_08', 'name': 'Lurdes Serafim',     'fraction': '3.º Direito',  'permilage': 88,
     'nif': '127143980', 'email': 'lurdesserafim@example.com',     'isAdmin': False,
     'rentByYear': {'2021': 3100, '2022': 3100, '2023': 3600, '2024': 3600, '2025': 3600, '2026': 3600}},
    {'id': 'cond_09', 'name': 'José Carlos Monteiro','fraction': '4.º Esquerdo', 'permilage': 125,
     'nif': '182258637', 'email': 'jcmonteiro@example.com',        'isAdmin': False,
     'rentByYear': {'2021': 4400, '2022': 4400, '2023': 5100, '2024': 5100, '2025': 5100, '2026': 5100}},
    {'id': 'cond_10', 'name': 'Vitor Barata',       'fraction': '4.º Direito',  'permilage': 87,
     'nif': '178132730', 'email': 'vitorbarata@example.com',       'isAdmin': False,
     'rentByYear': {'2021': 3100, '2022': 3100, '2023': 3600, '2024': 3600, '2025': 3600, '2026': 3600}},
]

# ──────────────────────────────────────────────────────────
# RÚBRICAS
# ──────────────────────────────────────────────────────────
RUBRICAS = [
    {'id': 'rub_edp',          'nome': 'EDP / Electricidade',  'cor': '#F59E0B', 'ativa': True, 'ordem': 1},
    {'id': 'rub_telefone',     'nome': 'Telefone',             'cor': '#06B6D4', 'ativa': False, 'ordem': 2,
     'descricao': 'Linha telefónica do prédio (terminou em 2022)'},
    {'id': 'rub_schindler',    'nome': 'Schindler / Elevador', 'cor': '#1E54C7', 'ativa': True, 'ordem': 3},
    {'id': 'rub_agua',         'nome': 'Água',                 'cor': '#3B82F6', 'ativa': True, 'ordem': 4},
    {'id': 'rub_limpeza',      'nome': 'Limpeza',              'cor': '#10B981', 'ativa': True, 'ordem': 5},
    {'id': 'rub_banc',         'nome': 'Despesas Bancárias',   'cor': '#6B7280', 'ativa': True, 'ordem': 6},
    {'id': 'rub_allianz',      'nome': 'Allianz / Seguros',    'cor': '#7C3AED', 'ativa': True, 'ordem': 7},
    {'id': 'rub_obras',        'nome': 'Obras / Intervenções', 'cor': '#DC2626', 'ativa': True, 'ordem': 8},
    {'id': 'rub_plano_schindler', 'nome': 'Plano Schindler', 'cor': '#0891B2', 'ativa': True, 'ordem': 9,
     'fornecedorDefault': 'Schindler', 'valorDefault': 42796,
     'descricao': 'Plano de pagamento faseado para reparação do elevador',
     'metodoPagamentoDefault': 'transferencia'},
    {'id': 'rub_outras',       'nome': 'Outras',               'cor': '#9CA3AF', 'ativa': True, 'ordem': 10},
]

# ──────────────────────────────────────────────────────────
# QUOTAS HISTÓRICAS · gera recibos por célula
# ──────────────────────────────────────────────────────────

MES_NAMES = ['Janeiro', 'Fevereiro', 'Março', 'Abril', 'Maio', 'Junho',
             'Julho', 'Agosto', 'Setembro', 'Outubro', 'Novembro', 'Dezembro']

receipts = []
receipt_counter_by_year = {}

def next_num(year):
    receipt_counter_by_year[year] = receipt_counter_by_year.get(year, 0) + 1
    return f"RCB H{receipt_counter_by_year[year]:03d}/{year}"

def parse_quotas_sheet(sheet_name, year):
    """Lê uma folha de Quotas YYYY e gera recibos.
    Linhas: 11-22 = Jan-Dez Receitas (em 2024-2026 começa na linha 10)."""
    ws = wb[sheet_name]
    # Encontrar onde começa "Receitas" (header linha) e onde começa Janeiro
    rows = list(ws.iter_rows(values_only=True))
    start_row = None
    for i, row in enumerate(rows):
        if row and len(row) > 0 and row[0] and str(row[0]).strip().startswith('Receitas'):
            start_row = i + 1  # Linha seguinte = Janeiro
            break
    if start_row is None:
        # Fallback: padrão
        start_row = 10 if year >= 2024 else 10  # 0-indexed
    # 12 meses
    for mes_idx in range(12):
        row = rows[start_row + mes_idx] if start_row + mes_idx < len(rows) else None
        if not row:
            continue
        # Coluna 0 deve ser o nome do mês (ou data)
        for col_idx, tenant_id in COL_TO_TENANT.items():
            val = row[col_idx] if col_idx < len(row) else None
            if val is None or val == '---' or val == '':
                continue
            try:
                val_eur = float(val)
            except (TypeError, ValueError):
                continue
            if val_eur <= 0:
                continue
            mes_num = mes_idx + 1
            valor_cent = int(round(val_eur * 100))
            data = f"{year}-{mes_num:02d}-15"  # convenção meio do mês
            tenant = next(t for t in TENANTS if t['id'] == tenant_id)
            quota_mensal = tenant['rentByYear'][str(year)]
            # Determinar coverage: assumimos cobre meses começando do próprio mês
            coverage = []
            restante = valor_cent
            ano_cov = year
            mes_cov = mes_num
            while restante >= quota_mensal:
                coverage.append({'year': ano_cov, 'month': mes_cov, 'valor_centimos': quota_mensal})
                restante -= quota_mensal
                mes_cov += 1
                if mes_cov > 12:
                    mes_cov = 1
                    ano_cov += 1
            if restante > 0:
                coverage.append({'year': ano_cov, 'month': mes_cov, 'valor_centimos': restante})
            receipts.append({
                'id': f'rcp_h_{year}_{mes_num:02d}_{tenant_id}',
                'tenantId': tenant_id,
                'valor_centimos': valor_cent,
                'date': data,
                'numero': next_num(year),
                'descricao': f'Quota {MES_NAMES[mes_idx]} {year} (histórico)',
                'metodoPagamento': 'transferencia',
                'coverage': coverage,
                'cobrancas': [],
                'criadoEm': datetime(year, mes_num, 15).timestamp() * 1000,
                'historico': True,
            })

for year, sheet_name in [(2021,'Quotas 2021'),(2022,'Quotas 2022'),(2023,'Quotas 2023'),
                          (2024,'Quotas 2024'),(2025,'Quotas 2025'),(2026,'Quotas 2026')]:
    parse_quotas_sheet(sheet_name, year)

print(f"Recibos gerados: {len(receipts)}")

# ──────────────────────────────────────────────────────────
# DESPESAS HISTÓRICAS · cada célula vira pagamento de despesa
# ──────────────────────────────────────────────────────────

pagamentosDespesa = []
despesa_counter = 0

def add_despesa(year, mes, val_eur, rubricaId, descricao):
    global despesa_counter
    if val_eur is None or val_eur == 0:
        return
    try:
        val_eur = float(val_eur)
    except (TypeError, ValueError):
        return
    if val_eur <= 0:
        return
    despesa_counter += 1
    pagamentosDespesa.append({
        'id': f'pd_h_{year}_{mes:02d}_{despesa_counter}',
        'rubricaId': rubricaId,
        'valor_centimos': int(round(val_eur * 100)),
        'date': f"{year}-{mes:02d}-15",
        'descricao': descricao,
        'metodoPagamento': 'transferencia',
        'fornecedor': '',
        'criadoEm': datetime(year, mes, 15).timestamp() * 1000,
        'historico': True,
    })

def parse_despesas(sheet_name, year):
    """Lê folha Despesas YYYY. Encontra mapa rúbricas → coluna."""
    ws = wb[sheet_name]
    rows = list(ws.iter_rows(values_only=True))
    # Encontrar linha do header (com EDP, ...)
    header_row = None
    for i, row in enumerate(rows):
        if row and 'EDP' in [str(c).upper() if c else '' for c in row]:
            header_row = i
            break
    if header_row is None:
        return
    header = rows[header_row]
    # Mapear nome → coluna
    col_to_rubrica = {}
    for col_idx, name in enumerate(header):
        if not name:
            continue
        n = str(name).strip().upper()
        if 'EDP' in n:           col_to_rubrica[col_idx] = 'rub_edp'
        elif 'TELEFONE' in n:    col_to_rubrica[col_idx] = 'rub_telefone'
        elif 'SCHINDLER' in n:   col_to_rubrica[col_idx] = 'rub_schindler'
        elif 'ÁGUA' in n or 'AGUA' in n: col_to_rubrica[col_idx] = 'rub_agua'
        elif 'LIMPEZA' in n:     col_to_rubrica[col_idx] = 'rub_limpeza'
        elif 'BANC' in n:        col_to_rubrica[col_idx] = 'rub_banc'
        elif 'ALLIANZ' in n:     col_to_rubrica[col_idx] = 'rub_allianz'
        elif 'INTERVENÇÕES' in n or 'INTERVENCOES' in n or 'OBRAS' in n: col_to_rubrica[col_idx] = 'rub_obras'
        elif 'PLANO PAGAMENTO' in n or 'PLANO PAG' in n: col_to_rubrica[col_idx] = 'rub_plano_schindler'
        elif 'OUTRAS' in n or 'OUTROS' in n: col_to_rubrica[col_idx] = 'rub_outras'

    # Linhas Janeiro a Dezembro (12 meses)
    for mes in range(1, 13):
        row_idx = header_row + mes
        if row_idx >= len(rows):
            break
        row = rows[row_idx]
        if not row:
            continue
        for col_idx, rubrica_id in col_to_rubrica.items():
            if col_idx < len(row):
                val = row[col_idx]
                add_despesa(year, mes, val, rubrica_id, f"{MES_NAMES[mes-1]} {year} (histórico)")

for year, sheet_name in [(2021,'Despesas 2021'),(2022,'Despesas 2022'),(2023,'Despesas 2023'),
                          (2024,'Despesas 2024'),(2025,'Despesas 2025'),(2026,'Despesas 2026')]:
    parse_despesas(sheet_name, year)

print(f"Despesas geradas: {len(pagamentosDespesa)}")

# ──────────────────────────────────────────────────────────
# PLANOS (Reparação Elevador 2023, Quotização Obras 2024, Reparação Elevador 2025)
# ──────────────────────────────────────────────────────────

planos = []
prestacoes = []

# Plano 1: Reparação Elevador 2023 (Ago-Out 2023)
plano_elev_2023 = {
    'id': 'plano_elev_2023',
    'nome': 'Reparação Elevador 2023',
    'descricao': 'Reparação extraordinária do elevador · 3 prestações',
    'tipo': 'valor_fixo',
    'total_centimos': 226500,
    'inicio': '2023-08-01',
    'fim': '2023-10-31',
    'estado': 'concluido',
    'criadoEm': datetime(2023, 8, 1).timestamp() * 1000,
    'numPrestacoes': 3,
    'historico': True,
}
planos.append(plano_elev_2023)

# Valores Reparação Elevador 2023 por fração (a partir do Excel · 3 mensais iguais)
valores_elev_2023 = {
    'cond_02': 69, 'cond_01': 60, 'cond_04': 66, 'cond_03': 90,
    'cond_06': 66, 'cond_05': 91, 'cond_08': 66, 'cond_07': 87,
    'cond_10': 66, 'cond_09': 94,
}
for tenant_id, valor_mensal in valores_elev_2023.items():
    for i, (ano, mes) in enumerate([(2023,8),(2023,9),(2023,10)]):
        prestacoes.append({
            'id': f'prest_elev2023_{tenant_id}_{i+1}',
            'planoId': 'plano_elev_2023',
            'tenantId': tenant_id,
            'numero': i+1,
            'valor_centimos': valor_mensal * 100,
            'dueDate': f"{ano}-{mes:02d}-15",
            'estado': 'pago',
            'pagoEm': datetime(ano, mes, 20).timestamp() * 1000,
            'historico': True,
        })

# Plano 2: Quotização Obras 2024 (Mar 2024 - Abr 2025) - concluído
plano_obras = {
    'id': 'plano_obras_2024',
    'nome': 'Quotização Obras 2024',
    'descricao': 'Quotização extraordinária para obras · 14 prestações',
    'tipo': 'valor_fixo',
    'total_centimos': 2500000,
    'inicio': '2024-03-24',
    'fim': '2025-04-24',
    'estado': 'concluido',
    'criadoEm': datetime(2024, 3, 1).timestamp() * 1000,
    'numPrestacoes': 14,
    'historico': True,
}
planos.append(plano_obras)

# Valores Quotização Obras (Excel · mensal por fração)
valores_obras = {
    'cond_02': 18958,   # 189.58 €
    'cond_01': 16458,   # 164.58 €
    'cond_04': 18125,
    'cond_03': 24792,
    'cond_06': 18333,
    'cond_05': 25208,
    'cond_08': 18333,
    'cond_07': 23958,
    'cond_10': 18125,
    'cond_09': 26042,
}

# Lista de 14 datas (Mar 2024 - Abr 2025)
datas_obras = []
y, m = 2024, 3
for _ in range(14):
    datas_obras.append((y, m))
    m += 1
    if m > 12: m = 1; y += 1

for tenant_id, valor_cent in valores_obras.items():
    for i, (ano, mes) in enumerate(datas_obras):
        prestacoes.append({
            'id': f'prest_obras_{tenant_id}_{i+1}',
            'planoId': 'plano_obras_2024',
            'tenantId': tenant_id,
            'numero': i+1,
            'valor_centimos': valor_cent,
            'dueDate': f"{ano}-{mes:02d}-24",
            'estado': 'pago',
            'pagoEm': datetime(ano, mes, 24).timestamp() * 1000,
            'historico': True,
        })

# Plano 3: Reparação Elevador 2025 (Abr-Nov 2025) - EM CURSO
plano_elev_2025 = {
    'id': 'plano_elev_2025',
    'nome': 'Reparação Elevador 2025',
    'descricao': 'Reparação extraordinária do elevador 2025 · 8 prestações mensais',
    'tipo': 'valor_fixo',
    'total_centimos': 543617,
    'inicio': '2025-04-01',
    'fim': '2025-11-30',
    'estado': 'ativo',
    'criadoEm': datetime(2025, 4, 1).timestamp() * 1000,
    'numPrestacoes': 8,
    'historico': True,
}
planos.append(plano_elev_2025)

# Total por fração · Reparação Elevador 2025 (Excel)
quotas_total_elev2025 = {
    'cond_02': 49469,   # 494.69 €
    'cond_01': 42946,
    'cond_04': 47295,   # Sílvia · em atraso 190.95€ (Excel)
    'cond_03': 64690,
    'cond_06': 47838,
    'cond_05': 65778,
    'cond_08': 47838,
    'cond_07': 62516,
    'cond_10': 47295,
    'cond_09': 67952,
}

# Para a Sílvia, falta exata = 190.95€ (Excel "Exercicio 2025")
# Pago = total - falta = 472.95 - 190.95 = 282 €
# Modelar: 4 primeiras prestações pagas a 70.50€ (= 282/4)
#          4 últimas pendentes a 47.74€ (= 190.95/4 com ajuste 0.01)
FALTA_SILVIA = 19095  # cêntimos
PAGO_SILVIA  = 28200  # cêntimos

elev2025_datas = [(2025,m) for m in range(4,12)]  # Abr-Nov 2025
for tenant_id, total_cent in quotas_total_elev2025.items():
    is_silvia = tenant_id == 'cond_04'

    if is_silvia:
        # 4 primeiras pagas (somam 282€) · 4 últimas pendentes (somam 190.95€)
        valores_prest = [
            7050, 7050, 7050, 7050,                       # 4 × 70.50 = 282.00
            4774, 4774, 4774, 4773,                       # 4 × ≈47.74 = 190.95
        ]
        estados = ['pago']*4 + ['pendente']*4
    else:
        # 8 prestações iguais · todas pagas
        valor_mensal = round(total_cent / 8)
        # Última absorve o residual
        ultima = total_cent - valor_mensal * 7
        valores_prest = [valor_mensal]*7 + [ultima]
        estados = ['pago']*8

    for i, (ano, mes) in enumerate(elev2025_datas):
        prest = {
            'id': f'prest_elev2025_{tenant_id}_{i+1}',
            'planoId': 'plano_elev_2025',
            'tenantId': tenant_id,
            'numero': i+1,
            'valor_centimos': valores_prest[i],
            'dueDate': f"{ano}-{mes:02d}-15",
            'historico': True,
            'estado': estados[i],
        }
        if estados[i] == 'pago':
            prest['pagoEm'] = datetime(ano, mes, 20).timestamp() * 1000
        prestacoes.append(prest)

print(f"Planos gerados: {len(planos)}")
print(f"Prestações geradas: {len(prestacoes)}")

# ──────────────────────────────────────────────────────────
# ORÇAMENTO 2026 APROVADO
# ──────────────────────────────────────────────────────────
orcamento_2026 = {
    'id': 'orc_2026_v1',
    'ano': 2026,
    'versao': 1,
    'estado': 'aprovado',
    'criadoEm': datetime(2025, 12, 1).timestamp() * 1000,
    'aprovadoEm': datetime(2025, 12, 15).timestamp() * 1000,
    'aprovadoPor': 'Ricardo Cordeiro',
    'arredondamento_centimos': 100,
    'incrementoPercentual': 0,
    'quotas': [
        {'tenantId': 'cond_02', 'permilage': 91, 'valorMensalAnterior': 3700, 'valorMensalNovo': 3700},
        {'tenantId': 'cond_01', 'permilage': 79, 'valorMensalAnterior': 3200, 'valorMensalNovo': 3200},
        {'tenantId': 'cond_04', 'permilage': 87, 'valorMensalAnterior': 3600, 'valorMensalNovo': 3600},
        {'tenantId': 'cond_03', 'permilage': 119,'valorMensalAnterior': 4800, 'valorMensalNovo': 4800},
        {'tenantId': 'cond_06', 'permilage': 88, 'valorMensalAnterior': 3600, 'valorMensalNovo': 3600},
        {'tenantId': 'cond_05', 'permilage': 121,'valorMensalAnterior': 4900, 'valorMensalNovo': 4900},
        {'tenantId': 'cond_08', 'permilage': 88, 'valorMensalAnterior': 3600, 'valorMensalNovo': 3600},
        {'tenantId': 'cond_07', 'permilage': 115,'valorMensalAnterior': 4700, 'valorMensalNovo': 4700},
        {'tenantId': 'cond_10', 'permilage': 87, 'valorMensalAnterior': 3600, 'valorMensalNovo': 3600},
        {'tenantId': 'cond_09', 'permilage': 125,'valorMensalAnterior': 5100, 'valorMensalNovo': 5100},
    ],
    'despesas': [
        {'rubricaId': 'rub_edp',          'orcado_centimos': 85352,  'realizado_centimos_anoAnterior': 77593},
        {'rubricaId': 'rub_agua',         'orcado_centimos': 23416,  'realizado_centimos_anoAnterior': 21287},
        {'rubricaId': 'rub_schindler',    'orcado_centimos': 140968, 'realizado_centimos_anoAnterior': 128153},
        {'rubricaId': 'rub_allianz',      'orcado_centimos': 84031,  'realizado_centimos_anoAnterior': 76392},
        {'rubricaId': 'rub_banc',         'orcado_centimos': 9972,   'realizado_centimos_anoAnterior': 9972},
        {'rubricaId': 'rub_limpeza',      'orcado_centimos': 120000, 'realizado_centimos_anoAnterior': 116400},
        {'rubricaId': 'rub_outras',       'orcado_centimos': 10000,  'realizado_centimos_anoAnterior': 0},
        {'rubricaId': 'rub_obras',        'orcado_centimos': 324835, 'realizado_centimos_anoAnterior': 0},
        {'rubricaId': 'rub_plano_schindler', 'orcado_centimos': 602326, 'realizado_centimos_anoAnterior': 0},
    ],
}

# ──────────────────────────────────────────────────────────
# META + Saldos
#
# Estrutura: meta vai conter docs com id (será expandido em coleção).
#   - condominio  · dados da entidade
#   - config      · saldoInicial por ano + saldoConhecido (ancoragem BPI)
# ──────────────────────────────────────────────────────────
meta = {
    'condominio': {
        'nome': 'Administração Condomínio Av. Amália Rodrigues, 24',
        'morada': 'Av. Amália Rodrigues, 24',
        'codigoPostal': '2650-437',
        'localidade': 'Amadora',
        'nif': '901589381',
        'iban': 'PT50 0010 0000 5398 9510 0019 7',
        'banco': 'BPI',
        'email': '',
        'telefone': '',
    },
    'config': {
        # Saldo inicial de cada ano (em cêntimos) · base do cálculo
        'saldoInicial': {
            '2026': 321478,    # CO 2510.44 + Poup 704.34 (01-Jan-2026)
        },
        # Saldo real observado · ancoragem para detectar descalibração
        'saldoConhecido': {
            'data': '2026-05-25',
            'contaOrdem_centimos': 752178,    # 7521.78 €
            'contaPoupanca_centimos': 70434,  # 704.34 €
            'total_centimos': 822612,         # 8226.12 €
            'notas': 'BPI Net Empresas · posição integrada · ponto de ancoragem inicial',
            'registadoEm': int(datetime.now().timestamp() * 1000),
        },
    },
    # Plano Schindler · pagamentos faseados do condomínio à Schindler
    # Excel: Entrada 10% + 12 prestações mensais (Dez 2025 - Nov 2026) = 6023.26 €
    'planoSchindler': {
        'inicio': '2025-11-01',
        'fim': '2026-11-30',
        'fornecedor': 'Schindler',
        'descricao': 'Plano de pagamento faseado para reparação do elevador',
        'totalPrevisto_centimos': 602326,
        'rubricaId': 'rub_plano_schindler',
        'prestacoes': [
            {'data': '2025-11-01', 'valor_centimos': 62326, 'descricao': 'Entrada (10%)'},
            {'data': '2025-12-01', 'valor_centimos': 45000, 'descricao': 'Prestação 1 / 12 · Dez 2025'},
            {'data': '2026-01-01', 'valor_centimos': 45000, 'descricao': 'Prestação 2 / 12 · Jan 2026'},
            {'data': '2026-02-01', 'valor_centimos': 45000, 'descricao': 'Prestação 3 / 12 · Fev 2026'},
            {'data': '2026-03-01', 'valor_centimos': 45000, 'descricao': 'Prestação 4 / 12 · Mar 2026'},
            {'data': '2026-04-01', 'valor_centimos': 45000, 'descricao': 'Prestação 5 / 12 · Abr 2026'},
            {'data': '2026-05-01', 'valor_centimos': 45000, 'descricao': 'Prestação 6 / 12 · Mai 2026'},
            {'data': '2026-06-01', 'valor_centimos': 45000, 'descricao': 'Prestação 7 / 12 · Jun 2026'},
            {'data': '2026-07-01', 'valor_centimos': 45000, 'descricao': 'Prestação 8 / 12 · Jul 2026'},
            {'data': '2026-08-01', 'valor_centimos': 45000, 'descricao': 'Prestação 9 / 12 · Ago 2026'},
            {'data': '2026-09-01', 'valor_centimos': 45000, 'descricao': 'Prestação 10 / 12 · Set 2026'},
            {'data': '2026-10-01', 'valor_centimos': 45000, 'descricao': 'Prestação 11 / 12 · Out 2026'},
            {'data': '2026-11-01', 'valor_centimos': 45000, 'descricao': 'Prestação 12 / 12 · Nov 2026'},
        ],
    },
}

# ──────────────────────────────────────────────────────────
# OUTROS RECEBIMENTOS (Devolução Reabilita+ e Upgrade Videoporteiro 2026)
# ──────────────────────────────────────────────────────────
outrosRecebimentos = [
    {
        'id': 'or_reabilita_2026',
        'data': '2026-04-15',
        'descricao': 'Devolução Reabilita+',
        'valor_centimos': 651900,
        'metodoPagamento': 'transferencia',
        'origem': 'Reabilita+',
        'criadoEm': datetime(2026, 4, 15).timestamp() * 1000,
        'historico': True,
    },
    {
        'id': 'or_videoporteiro_2026',
        'data': '2026-03-20',
        'descricao': 'Upgrade Videoporteiro · contribuições dos condóminos',
        'valor_centimos': 14500,
        'metodoPagamento': 'transferencia',
        'origem': 'Condóminos',
        'criadoEm': datetime(2026, 3, 20).timestamp() * 1000,
        'historico': True,
    },
]

# ──────────────────────────────────────────────────────────
# Output
# ──────────────────────────────────────────────────────────

snapshot = {
    'meta': meta,
    'tenants': TENANTS,
    'rubricas': RUBRICAS,
    'receipts': receipts,
    'pagamentosDespesa': pagamentosDespesa,
    'planos': planos,
    'prestacoes': prestacoes,
    'orcamentos': [orcamento_2026],
    'outrosRecebimentos': outrosRecebimentos,
    'comunicacoes': [],
    '__importInfo': {
        'geradoEm': datetime.now().isoformat(),
        'origem': 'Backup Excel Condomínio AR24 v3.0',
        'periodo': '2021-2026',
        'contagens': {
            'tenants': len(TENANTS),
            'rubricas': len(RUBRICAS),
            'receipts': len(receipts),
            'despesas': len(pagamentosDespesa),
            'planos': len(planos),
            'prestacoes': len(prestacoes),
            'outrosRecebimentos': len(outrosRecebimentos),
        }
    }
}

with open(OUT, 'w', encoding='utf-8') as f:
    json.dump(snapshot, f, ensure_ascii=False, indent=2)

print(f"\n✅ Snapshot gerado em: {OUT}")
print(f"Tamanho: {os.path.getsize(OUT) / 1024:.1f} KB")
print()
print("Sumário:")
for k, v in snapshot['__importInfo']['contagens'].items():
    print(f"  • {k}: {v}")
